"""Excel workbook generation for Wayback snapshot exports.

This module converts raw snapshot URLs into a user-friendly workbook with
separate HTTPS/HTTP sheets and a summary sheet. GUI code should call these
functions instead of manipulating pandas or openpyxl directly.
"""

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from modules.paths import PROCESSED_DIR
from modules.wayback import convert_to_http, extract_date_from_link, format_date_columns


def save_snapshots_to_excel(snapshots: list[str], filename: str | Path) -> Path:
    """Save Wayback snapshot links to a formatted Excel workbook.

    Args:
        snapshots: List of Wayback snapshot URL strings.
        filename: Destination path as a string or ``Path`` object.

    Returns:
        ``Path`` to the saved workbook.

    Raises:
        ValueError: If ``snapshots`` is empty.
    """
    # If there are no captures, stop before creating an empty workbook that
    # would look successful but contain no useful data.
    if not snapshots:
        raise ValueError("No snapshots were provided.")

    # Normalize the output path and create its parent directory before writing.
    output_path = Path(filename)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Build one row per snapshot, including a parsed datetime for later
    # formatting and summary calculations.
    data = [
        (index + 1, snapshot, extract_date_from_link(snapshot))
        for index, snapshot in enumerate(snapshots)
    ]
    df = pd.DataFrame(data, columns=["Capture", "Capture Link", "Date"])

    # Expand formatted date values into human-readable Excel columns.
    formatted_dates = df["Date"].apply(format_date_columns).apply(pd.Series)

    # Keep the original capture links on the primary worksheet.
    df_with_dates = pd.concat([df[["Capture", "Capture Link"]], formatted_dates], axis=1)

    # Create a second sheet with HTTP-converted links for workflows that need
    # legacy HTTP capture URLs.
    df_http = df.copy()
    df_http["Capture Link"] = df_http["Capture Link"].apply(convert_to_http)
    df_with_dates_http = pd.concat(
        [df_http[["Capture", "Capture Link"]], formatted_dates],
        axis=1,
    )

    # Summarize the capture set so users can quickly see the time span covered.
    summary_df = pd.DataFrame(
        {
            "Total Captures": [len(snapshots)],
            "First Capture Date": [df["Date"].min().strftime("%B %d, %Y %I:%M:%S %p")],
            "Last Capture Date": [df["Date"].max().strftime("%B %d, %Y %I:%M:%S %p")],
        }
    )

    # Write all sheets in one writer context so pandas handles workbook creation
    # before openpyxl does the styling pass.
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_with_dates.to_excel(writer, sheet_name="Snapshots", index=False)
        df_with_dates_http.to_excel(writer, sheet_name="Snapshots_HTTP", index=False)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

    # Apply column widths and alignment after pandas writes raw data.
    _format_workbook(output_path)
    return output_path


def snapshot_excel_path(domain: str) -> Path:
    """Build the default snapshot workbook path for a domain.

    Args:
        domain: Domain string such as ``example.com``.

    Returns:
        ``Path`` under ``reports/processed`` for the domain's snapshot workbook.
    """
    # Keep output naming consistent across GUI and future command-line callers.
    return PROCESSED_DIR / f"{domain}_snapshots.xlsx"


def _format_workbook(filename: Path) -> None:
    """Apply formatting to every worksheet in an Excel workbook.

    Args:
        filename: ``Path`` to an existing workbook.

    Returns:
        None. The workbook is modified and saved in place.
    """
    # Reopen the workbook with openpyxl because it provides direct cell styling.
    workbook = load_workbook(filename)

    # Apply the same readable formatting to every sheet in the workbook.
    for worksheet in workbook.worksheets:
        _autosize_and_align(worksheet)
    workbook.save(filename)


def _autosize_and_align(worksheet) -> None:
    """Autosize columns and center-align cells on one worksheet.

    Args:
        worksheet: openpyxl worksheet object to format.

    Returns:
        None. The worksheet object is modified in memory.
    """
    # Iterate by column so each width is based on the longest value in that
    # column.
    for column in worksheet.columns:
        column_letter = get_column_letter(column[0].column)
        max_length = max(len(str(cell.value or "")) for cell in column)
        worksheet.column_dimensions[column_letter].width = max_length + 2

        # Center and wrap every cell so long snapshot URLs remain readable.
        for cell in column:
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

    # Let Excel choose row heights automatically after wrapping text.
    for row in worksheet.iter_rows():
        worksheet.row_dimensions[row[0].row].height = None
