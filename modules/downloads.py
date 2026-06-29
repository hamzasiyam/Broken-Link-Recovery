"""Download and wget logging utilities for archive recovery workflows.

This module isolates all subprocess and spreadsheet-download behavior from the
Tkinter interface. GUI modules can call these functions to validate wget,
download archive snapshots, and write log workbooks without knowing the command
details.
"""

import shutil
import subprocess
import time
from pathlib import Path
from typing import Callable, Iterable

import openpyxl
import pandas as pd

from modules.paths import DOWNLOADED_FILES_DIR, PROCESSED_DIR, WGET_EXE


ProgressCallback = Callable[[str], None]


def resolve_wget_executable() -> str:
    """Choose the wget executable the application should use.

    Args:
        None.

    Returns:
        Path string for a system ``wget``, the bundled ``wget.exe``, or the
        fallback command name ``"wget"``.
    """
    # Prefer a system installation because it is most likely to match the user's
    # platform and PATH configuration.
    system_wget = shutil.which("wget")
    if system_wget:
        return system_wget

    # If no system wget exists, use the bundled Windows executable when present.
    if WGET_EXE.exists():
        return str(WGET_EXE)

    # Return the plain command name as a final fallback; check_wget will report a
    # useful error if the command cannot actually run.
    return "wget"


def check_wget(wget_path: str | None = None) -> tuple[bool, str]:
    """Verify that wget can run.

    Args:
        wget_path: Optional executable path string. If omitted, the resolver
            selects the best available wget executable.

    Returns:
        Tuple ``(is_available, message)`` where ``is_available`` is a boolean
        and ``message`` is either the version line or an error message.
    """
    # Use the provided path for dependency injection, otherwise resolve the
    # executable from the environment.
    executable = wget_path or resolve_wget_executable()
    try:
        # ``wget --version`` is a lightweight command that confirms the binary
        # can be started.
        result = subprocess.run(
            [executable, "--version"],
            capture_output=True,
            text=True,
            check=False,
        )
    except OSError as exc:
        # If the operating system cannot execute the command, return a readable
        # message instead of raising into the GUI callback.
        return False, str(exc)

    # If wget starts but reports an error, surface stderr or a generic fallback.
    if result.returncode != 0:
        return False, result.stderr.strip() or "wget is not available."

    # Return the first version line so users can confirm which executable is in
    # use.
    first_line = result.stdout.splitlines()[0] if result.stdout else "wget is available."
    return True, first_line


def sanitize_filename(filename: str) -> str:
    """Replace unsafe filename characters with underscores.

    Args:
        filename: Proposed filename or folder name string.

    Returns:
        Sanitized string containing only alphanumeric characters, spaces, dots,
        and underscores.
    """
    # Keep readable characters and replace anything risky for Windows/macOS/Linux
    # paths.
    return "".join(
        character if character.isalnum() or character in (" ", ".", "_") else "_"
        for character in str(filename)
    )


class WgetDownloader:
    """Stateful wrapper around wget subprocess downloads.

    Args:
        wget_path: Optional path string to a wget executable.

    Returns:
        A downloader instance that tracks its active subprocess so the GUI can
        terminate it when the window closes.
    """

    def __init__(self, wget_path: str | None = None) -> None:
        """Initialize a downloader and resolve the wget executable.

        Args:
            wget_path: Optional path string to a wget executable.

        Returns:
            None. The constructor stores executable path and process state.
        """
        # Resolve the executable once so all downloads in the same GUI session
        # use the same binary.
        self.wget_path = wget_path or resolve_wget_executable()

        # Track the active process so ``terminate`` can stop long downloads.
        self.current_process: subprocess.Popen | None = None

    def download_archive_snapshot(
        self,
        url: str,
        target_dir: str | Path,
        capture: str,
        date_str: str,
        time_str: str,
        progress_callback: ProgressCallback = print,
    ) -> bool:
        """Download one archived snapshot into a dated output folder.

        Args:
            url: Snapshot URL string to download with wget.
            target_dir: Parent directory path where the capture folder is
                created.
            capture: Capture identifier from the spreadsheet.
            date_str: Human-readable capture date string.
            time_str: Human-readable capture time string.
            progress_callback: Callable that accepts status strings. Defaults
                to ``print``.

        Returns:
            ``True`` when wget exits with status ``0``; otherwise ``False``.
        """
        # Colons are not valid in Windows folder names, so replace them before
        # constructing the capture folder.
        safe_time = str(time_str).replace(":", "_")
        folder_name = sanitize_filename(f"Capture_{capture}_Date_{date_str}_Time_{safe_time}")
        final_target_dir = Path(target_dir) / folder_name
        final_target_dir.mkdir(parents=True, exist_ok=True)
        progress_callback(f"Created directory: {final_target_dir}")

        # Build the wget command as a list to avoid shell quoting issues and keep
        # arguments explicit for future maintainers.
        command = [
            self.wget_path,
            "--mirror",
            "--convert-links",
            "--adjust-extension",
            "--page-requisites",
            "--no-parent",
            str(url),
            "-P",
            str(final_target_dir),
        ]
        progress_callback(f"Running command: {' '.join(command)}")

        try:
            # Start wget and merge stderr into stdout so the callback sees one
            # ordered stream of progress messages.
            self.current_process = subprocess.Popen(
                command,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
            )

            # If stdout is available, forward each line to the caller's progress
            # handler as wget runs.
            if self.current_process.stdout:
                for line in self.current_process.stdout:
                    progress_callback(line.strip())

            self.current_process.wait()

            # If wget failed, report failure without raising so the spreadsheet
            # processor can continue with later rows.
            if self.current_process.returncode != 0:
                progress_callback(f"Error downloading {url}.")
                return False

            # If wget succeeded, report the completed destination folder.
            progress_callback(f"Successfully downloaded: {url} to {final_target_dir}")
            return True
        except OSError as exc:
            # OS errors usually mean the executable could not start or a path is
            # invalid. Return False so the GUI can keep control.
            progress_callback(f"Error downloading {url}: {exc}")
            return False
        finally:
            # Clear process state whether the command succeeds, fails, or raises.
            self.current_process = None

    def terminate(self) -> None:
        """Terminate the active wget subprocess, if one is running.

        Args:
            None.

        Returns:
            None. The active process is terminated and cleared when present.
        """
        # If no process is active, there is nothing to stop.
        if self.current_process is not None:
            # Terminate the active process so closing the GUI does not leave wget
            # running in the background.
            self.current_process.terminate()
            self.current_process = None


def process_spreadsheet(
    file_path: str | Path,
    sheet_name: str,
    output_dir: str | Path,
    delay_seconds: int,
    downloader: WgetDownloader | None = None,
    progress_callback: ProgressCallback = print,
) -> Path:
    """Download every snapshot listed in a spreadsheet sheet.

    Args:
        file_path: Path string or ``Path`` to an Excel workbook.
        sheet_name: Name of the worksheet to process.
        output_dir: Parent directory where downloaded contents should be saved.
        delay_seconds: Integer pause between downloads.
        downloader: Optional ``WgetDownloader`` instance. If omitted, one is
            created.
        progress_callback: Callable that receives progress strings.

    Returns:
        ``Path`` to the domain-specific output directory.

    Raises:
        ValueError: If required spreadsheet columns are missing.
    """
    # Normalize the spreadsheet path before reading and deriving the output name.
    spreadsheet_path = Path(file_path)
    df = pd.read_excel(spreadsheet_path, sheet_name=sheet_name)

    # These column names are produced by the snapshot Excel workflow and are
    # required for naming folders and downloading capture links.
    required_columns = [
        "Capture",
        "Capture Link",
        "Month Day, Year",
        "Hour, Minute, Second AM/PM",
    ]
    missing_columns = [column for column in required_columns if column not in df.columns]

    # If the user picked the wrong sheet or workbook, explain exactly which
    # columns are missing.
    if missing_columns:
        missing = ", ".join(missing_columns)
        raise ValueError(f"The spreadsheet is missing required columns: {missing}")

    # Use the domain prefix from the workbook filename to group downloads for
    # that site.
    domain = spreadsheet_path.name.split("_")[0]
    specific_output_dir = Path(output_dir) / f"{domain}_snapshots"
    specific_output_dir.mkdir(parents=True, exist_ok=True)

    # Reuse the caller-provided downloader when possible so GUI close handling
    # can terminate the active process.
    active_downloader = downloader or WgetDownloader()

    # Iterate over each spreadsheet row and download the capture link listed in
    # that row.
    for _, row in df.iterrows():
        capture = row["Capture"]
        url = row["Capture Link"]
        date_str = row["Month Day, Year"]
        time_str = row["Hour, Minute, Second AM/PM"]
        progress_callback(f"Downloading URL: {url}")
        success = active_downloader.download_archive_snapshot(
            url,
            specific_output_dir,
            capture,
            date_str,
            time_str,
            progress_callback,
        )

        # If one URL fails, log the failure and continue with the next row.
        if not success:
            progress_callback(f"Failed to download: {url}")

        # Pause between downloads to avoid hammering remote services.
        time.sleep(delay_seconds)

    return specific_output_dir


def run_wget_and_log_to_excel(command: Iterable[str], output_excel: str | Path) -> int:
    """Run a wget command and capture stdout/stderr into an Excel log.

    Args:
        command: Iterable of command argument strings, including the executable.
        output_excel: Path string or ``Path`` where the log workbook is saved.

    Returns:
        Integer process return code from wget.
    """
    # Convert the command to a list so subprocess receives a stable argument
    # sequence even if the caller passed another iterable type.
    process = subprocess.Popen(
        list(command),
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
    )
    stdout, stderr = process.communicate()

    # Create a simple workbook with timestamp, event type, and output details.
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Wget Log"
    sheet["A1"] = "Timestamp"
    sheet["B1"] = "Event"
    sheet["C1"] = "Details"

    row_index = 2
    from datetime import datetime

    # Use one timestamp for all lines from this run so log rows are grouped.
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Write stdout as "Output" rows and stderr as "Error" rows.
    for event_name, output in (("Output", stdout), ("Error", stderr)):
        for line in output.splitlines():
            sheet[f"A{row_index}"] = timestamp
            sheet[f"B{row_index}"] = event_name
            sheet[f"C{row_index}"] = line
            row_index += 1

    # Ensure the log directory exists before saving the workbook.
    output_path = Path(output_excel)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return process.returncode


def download_range_from_excel(
    excel_file: str | Path,
    start_row: int,
    end_row: int,
    download_dir: str | Path = DOWNLOADED_FILES_DIR,
    log_dir: str | Path = PROCESSED_DIR,
) -> list[Path]:
    """Download capture links from a row range in an Excel workbook.

    Args:
        excel_file: Path string or ``Path`` to an Excel workbook.
        start_row: First 1-based worksheet row number to process.
        end_row: Last 1-based worksheet row number to process.
        download_dir: Directory where wget should place downloaded files.
        log_dir: Directory where wget log workbooks should be saved.

    Returns:
        List of ``Path`` objects for the generated wget log workbooks.
    """
    # Open the workbook and use the active sheet because the legacy workflow did
    # not expose sheet selection for this range-based tool.
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active
    wget_path = resolve_wget_executable()
    log_paths: list[Path] = []

    # Process the inclusive row range selected by the user.
    for row in range(start_row, end_row + 1):
        capture_link = sheet.cell(row=row, column=2).value

        # If the row has no capture link in column B, skip it silently.
        if not capture_link:
            continue

        # Build the legacy recursive wget command used by this workflow.
        command = [
            wget_path,
            "-r",
            "-l",
            "inf",
            "-np",
            "-nH",
            "--cut-dirs=1",
            "-P",
            str(download_dir),
            "--convert-links",
            "-e",
            "robots=off",
            "-U",
            "Mozilla",
            str(capture_link),
        ]
        output_excel = Path(log_dir) / f"wget_log_{row}.xlsx"

        # Run wget for this row and save its output to a per-row log workbook.
        run_wget_and_log_to_excel(command, output_excel)
        log_paths.append(output_excel)

    return log_paths
