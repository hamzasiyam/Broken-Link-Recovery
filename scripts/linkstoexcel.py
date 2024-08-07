import subprocess
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import os
import tkinter as tk
from tkinter import filedialog, messagebox

def get_snapshots(url):
    result = subprocess.run(['waybackpack', url, '--list'], capture_output=True, text=True)
    if result.returncode != 0:
        print("Error fetching snapshots")
        print(result.stderr)
        return []
    snapshots = result.stdout.strip().split('\n')
    return snapshots

def extract_date_from_link(link):
    timestamp = link.split('/')[4]
    date = datetime.strptime(timestamp, '%Y%m%d%H%M%S')
    return date

def format_date_columns(date):
    return {
        'Month Day, Year': date.strftime('%B %d, %Y'),
        'Hour, Minute, Second AM/PM': date.strftime('%I:%M:%S %p')
    }

def convert_to_http(link):
    return link.replace('https://', 'http://')

def save_to_excel(snapshots, filename):
    data = [(i + 1, snap, extract_date_from_link(snap)) for i, snap in enumerate(snapshots)]
    df = pd.DataFrame(data, columns=['Capture', 'Capture Link', 'Date'])
    formatted_dates = df['Date'].apply(format_date_columns).apply(pd.Series)
    df_with_dates = pd.concat([df[['Capture', 'Capture Link']], formatted_dates], axis=1)
    
    df_http = df.copy()
    df_http['Capture Link'] = df_http['Capture Link'].apply(convert_to_http)
    df_with_dates_http = pd.concat([df_http[['Capture', 'Capture Link']], formatted_dates], axis=1)
    
    # Capture the date columns separately for summary
    first_capture_date = df['Date'].min()
    last_capture_date = df['Date'].max()
    
    summary = {
        'Total Captures': [len(snapshots)],
        'First Capture Date': [first_capture_date.strftime('%B %d, %Y %I:%M:%S %p')],
        'Last Capture Date': [last_capture_date.strftime('%B %d, %Y %I:%M:%S %p')]
    }
    summary_df = pd.DataFrame(summary)
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df_with_dates.to_excel(writer, sheet_name='Snapshots', index=False)
        df_with_dates_http.to_excel(writer, sheet_name='Snapshots_HTTP', index=False)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
    
    wb = load_workbook(filename)
    ws_snapshots = wb['Snapshots']
    ws_snapshots_http = wb['Snapshots_HTTP']
    
    for ws in [ws_snapshots, ws_snapshots_http]:
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[col_letter].width = adjusted_width
            for cell in col:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws_summary = wb['Summary']
    for col in ws_summary.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws_summary.column_dimensions[col_letter].width = adjusted_width

    for row in ws_snapshots.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws_snapshots.row_dimensions[row[0].row].height = None

    for row in ws_snapshots_http.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws_snapshots_http.row_dimensions[row[0].row].height = None
    
    for row in ws_summary.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws_summary.row_dimensions[row[0].row].height = None
    
    wb.save(filename)

def process_snapshots():
    url = entry_url.get()
    if not url:
        messagebox.showerror("Error", "Please enter a URL.")
        return
    snapshots = get_snapshots(url)
    if snapshots:
        domain = url.split('//')[-1].split('/')[0]
        filename = f'reports/processed/{domain}_snapshots.xlsx'
        os.makedirs(os.path.dirname(filename), exist_ok=True)
        save_to_excel(snapshots, filename)
        messagebox.showinfo("Success", f'Snapshots and summary saved to {filename}')
    else:
        messagebox.showerror("Error", "No snapshots found or there was an error.")

app = tk.Tk()
app.title("Wayback Machine Snapshot Exporter")

frame = tk.Frame(app, padx=10, pady=10)
frame.pack(padx=10, pady=10)

label_url = tk.Label(frame, text="Enter URL:")
label_url.grid(row=0, column=0, pady=5)

entry_url = tk.Entry(frame, width=50)
entry_url.grid(row=0, column=1, pady=5)

button_process = tk.Button(frame, text="Process Snapshots", command=process_snapshots)
button_process.grid(row=1, columnspan=2, pady=10)

app.mainloop()
