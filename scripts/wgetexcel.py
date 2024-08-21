import tkinter as tk
from tkinter import filedialog, messagebox
import subprocess
import openpyxl
from datetime import datetime
import os

def run_wget_and_log_to_excel(wget_command, output_excel):
    # Run the wget command and capture the output
    process = subprocess.Popen(wget_command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, universal_newlines=True)
    stdout, stderr = process.communicate()
    
    # Create or open the Excel workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Wget Log"
    
    # Set headers
    sheet['A1'] = "Timestamp"
    sheet['B1'] = "Event"
    sheet['C1'] = "Details"
    
    # Write the wget output to the Excel sheet
    row = 2
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    if stdout:
        for line in stdout.splitlines():
            sheet[f'A{row}'] = timestamp
            sheet[f'B{row}'] = "Output"
            sheet[f'C{row}'] = line
            row += 1
            
    if stderr:
        for line in stderr.splitlines():
            sheet[f'A{row}'] = timestamp
            sheet[f'B{row}'] = "Error"
            sheet[f'C{row}'] = line
            row += 1
    
    # Save the workbook
    workbook.save(output_excel)
    print(f"Log saved to {output_excel}")

def select_excel_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        excel_file_entry.delete(0, tk.END)
        excel_file_entry.insert(0, file_path)

def download_range():
    excel_file = excel_file_entry.get()
    start_row = int(start_row_entry.get())
    end_row = int(end_row_entry.get())

    if not os.path.exists(excel_file):
        messagebox.showerror("Error", "Please select a valid Excel file.")
        return

    # Load the workbook and select the active sheet
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active
    
    for row in range(start_row, end_row + 1):
        capture_link = sheet.cell(row=row, column=2).value
        timestamp = sheet.cell(row=row, column=3).value + " " + sheet.cell(row=row, column=4).value

        wget_command = f'C:\\Windows\\System32\\wget.exe -r -l inf -np -nH --cut-dirs=1 -P ./downloaded_files --convert-links -e robots=off -U Mozilla {capture_link}'
        output_excel = f'wget_log_{row}.xlsx'
        
        run_wget_and_log_to_excel(wget_command, output_excel)
    
    messagebox.showinfo("Success", "Download and logging completed!")

# GUI Setup
root = tk.Tk()
root.title("Wget Downloader")

tk.Label(root, text="Excel File:").grid(row=0, column=0, padx=10, pady=10)
excel_file_entry = tk.Entry(root, width=50)
excel_file_entry.grid(row=0, column=1, padx=10, pady=10)
tk.Button(root, text="Browse", command=select_excel_file).grid(row=0, column=2, padx=10, pady=10)

tk.Label(root, text="Start Row:").grid(row=1, column=0, padx=10, pady=10)
start_row_entry = tk.Entry(root, width=10)
start_row_entry.grid(row=1, column=1, padx=10, pady=10, sticky="w")

tk.Label(root, text="End Row:").grid(row=2, column=0, padx=10, pady=10)
end_row_entry = tk.Entry(root, width=10)
end_row_entry.grid(row=2, column=1, padx=10, pady=10, sticky="w")

tk.Button(root, text="Download", command=download_range).grid(row=3, column=1, padx=10, pady=20)

root.mainloop()
